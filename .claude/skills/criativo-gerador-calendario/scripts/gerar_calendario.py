import calendar
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def create_excel_calendar(output_path, year=2026):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Calendario {year}"
    
    font_bold = Font(bold=True, size=11)
    font_title = Font(bold=True, size=14, color="FFFFFF")
    
    fill_header = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_month = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin = Side(border_style="thin", color="000000")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for i in range(1, 15):
        ws.column_dimensions[chr(64+i)].width = 20
        
    ws.merge_cells('A1:N1')
    ws['A1'] = str(year)
    ws['A1'].font = font_title
    ws['A1'].fill = fill_header
    ws['A1'].alignment = alignment_center
    
    days = ["DOMINGO", "SEGUNDA-FEIRA", "TERCA-FEIRA", "QUARTA-FEIRA", "QUINTA-FEIRA", "SEXTA-FEIRA", "SABADO"]
    col = 1
    for day in days:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        cell = ws.cell(row=3, column=col)
        cell.value = day
        cell.font = font_bold
        cell.alignment = alignment_center
        cell.border = box_border
        ws.cell(row=3, column=col+1).border = box_border
        col += 2
        
    current_row = 4
    cal = calendar.Calendar(firstweekday=6)
    months = ["JANEIRO", "FEVEREIRO", "MARCO", "ABRIL", "MAIO", "JUNHO", 
              "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]
              
    for month_idx, month_name in enumerate(months):
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        cell = ws.cell(row=current_row, column=1)
        cell.value = month_name
        cell.font = font_title
        cell.fill = fill_month
        cell.alignment = alignment_center
        
        for c in range(1, 15):
            ws.cell(row=current_row, column=c).border = box_border

        current_row += 1
        
        weeks = cal.monthdayscalendar(int(year), month_idx + 1)
        for week in weeks:
            col = 1
            for day in week:
                ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+1)
                cell = ws.cell(row=current_row, column=col)
                if day != 0:
                    cell.value = day
                cell.font = font_bold
                cell.alignment = alignment_center
                
                cell.border = box_border
                ws.cell(row=current_row, column=col+1).border = box_border
                
                ws.merge_cells(start_row=current_row+1, start_column=col, end_row=current_row+6, end_column=col+1)
                content_cell = ws.cell(row=current_row+1, column=col)
                content_cell.alignment = alignment_top
                
                for r in range(current_row+1, current_row+7):
                    ws.cell(row=r, column=col).border = box_border
                    ws.cell(row=r, column=col+1).border = box_border
                
                col += 2
            current_row += 7

    wb.save(output_path)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        out = sys.argv[1]
        yr = sys.argv[2] if len(sys.argv) > 2 else 2026
        create_excel_calendar(out, int(yr))
        print("Excel gerado com sucesso!")
    else:
        print("Uso: python gerar_calendario.py <caminho_saida.xlsx> [ano]")
